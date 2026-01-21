# ===============================
# EXCEL EXPORT
# ===============================
"""
Handles Excel file creation and export with separate worksheets for ambiguous entries
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from system import config_loader

def filter_out_original_terms(current_nouns, original_nouns):
    """
    Filter out terms that were already in the original glossary.
    
    Args:
        current_nouns (list): Current list of nouns after processing
        original_nouns (list): Original list of nouns before processing
        
    Returns:
        list: Filtered list containing only new terms
    """
    if not original_nouns:
        return current_nouns
    
    # Create a set of hangul from original glossary for quick lookup
    original_hanguls = {noun['hangul'] for noun in original_nouns}
    
    # Filter out terms that exist in the original glossary
    filtered_nouns = []
    for noun in current_nouns:
        if noun['hangul'] not in original_hanguls:
            filtered_nouns.append(noun)
    
    return filtered_nouns

def export_to_excel(excel_data, categories):
    """
    Export data to Excel files as specified:
    1. {output_excel}_master.xlsx - 2 worksheets: Full Glossary and ambiguous terms (if ambiguous exists)
    2. {output_excel}.xlsx - Worksheets for each category (clear only) + ambiguous terms worksheet (if ambiguous exists)
    
    Args:
        excel_data: List of noun dictionaries
        categories: List of category names
        
    Returns:
        Boolean success status
    """
    try:
        print(f"\nExporting to Excel files...")
        
        # Define the fixed workbook order
        FIXED_CATEGORY_ORDER = [
            'character names',
            'character titles', 
            'skills and techniques',
            'locations and organizations',
            'item names',
            'misc'
        ]
        
        # Convert to DataFrame
        df = pd.DataFrame(excel_data)
        
        # Ensure ambiguous column exists (default to False)
        if 'ambiguous' not in df.columns:
            df['ambiguous'] = False
        else:
            df['ambiguous'] = df['ambiguous'].fillna(False).astype(bool)
        
        # Ensure frequency is integer
        if 'frequency' in df.columns:
            df['frequency'] = df['frequency'].fillna(0).astype(int)

        if config_loader.SIMPLIFIED_CHINESE_CONVERSION:
            MASTER_COLUMNS = ['hangul', 'hanja', 'chinese', 'english', 'category', 'frequency']
            MAIN_COLUMNS   = ['hangul', 'hanja', 'chinese', 'english', 'frequency']
        else:
            MASTER_COLUMNS = ['hangul', 'hanja', 'english', 'category', 'frequency']
            MAIN_COLUMNS   = ['hangul', 'hanja', 'english', 'frequency']
        
        # Separate clear and ambiguous entries
        clear_df = df[df['ambiguous'] == False].copy()
        ambiguous_df = df[df['ambiguous'] == True].copy()
        
        print(f"  Total entries: {len(df)}")
        print(f"  Full Glossary entries: {len(clear_df)}")
        print(f"  ambiguous terms entries: {len(ambiguous_df)}")
        
        # Sort function for both DataFrames
        def sort_dataframe(df):
            """Sort DataFrame by hangul length (descending) then frequency (descending)"""
            if len(df) == 0:
                return df
            
            # Calculate hangul length
            df = df.copy()
            df['hangul_len'] = df['hangul'].str.len()
            
            # Sort by length then frequency
            sorted_df = df.sort_values(['hangul_len', 'frequency'], ascending=[False, False])
            
            # Drop the temporary column
            sorted_df = sorted_df.drop(columns=['hangul_len'])
            
            return sorted_df
        
        # Sort both DataFrames
        clear_df = sort_dataframe(clear_df)
        ambiguous_df = sort_dataframe(ambiguous_df)
        
        # 1. Create MASTER Excel file ({output_excel}_master.xlsx)
        master_filename = config_loader.OUTPUT_EXCEL.replace('.xlsx', '_master.xlsx')
        if master_filename == config_loader.OUTPUT_EXCEL:  # In case there's no .xlsx extension
            master_filename = config_loader.OUTPUT_EXCEL + '_master.xlsx'
        
        print(f"\nCreating master Excel file: {master_filename}")
        
        with pd.ExcelWriter(master_filename, engine='openpyxl') as writer:
            # Write Full Glossary entries sheet (always create if there are clear entries)
            if not clear_df.empty:
                clear_master_df = clear_df[[c for c in MASTER_COLUMNS if c in clear_df.columns]]
                clear_master_df.to_excel(writer, sheet_name='full glossary', index=False)
                print(f"  - Full Glossary entries: {len(clear_master_df)} rows")
            else:
                print(f"  - No clear entries to write")
            
            # Write ambiguous terms entries sheet (only if there are ambiguous entries)
            if not ambiguous_df.empty:
                ambiguous_master_df = ambiguous_df[[c for c in MASTER_COLUMNS if c in ambiguous_df.columns]]
                ambiguous_master_df.to_excel(writer, sheet_name='ambiguous terms', index=False)
                print(f"  - ambiguous terms entries: {len(ambiguous_master_df)} rows")
            
            # Auto-adjust column widths for all sheets
            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✓ Master file created successfully")
        
        # 2. Create MAIN Excel file ({output_excel}.xlsx) with fixed category order
        main_filename = config_loader.OUTPUT_EXCEL
        print(f"\nCreating main Excel file: {main_filename}")
        
        with pd.ExcelWriter(main_filename, engine='openpyxl') as writer:
            # For clear entries: split by category and create worksheets in fixed order
            category_count = 0
            if not clear_df.empty and 'category' in clear_df.columns:
                # Create worksheets in the fixed order
                for category in FIXED_CATEGORY_ORDER:
                    # Filter by category (case-insensitive match)
                    category_mask = clear_df['category'].str.lower() == category.lower()
                    category_df = clear_df[category_mask].copy()
                    
                    # Only create worksheet if there are entries
                    if not category_df.empty:
                        # Select columns using MAIN_COLUMNS
                        category_df = category_df[[c for c in MAIN_COLUMNS if c in category_df.columns]]
                        
                        # Use original category name from data for sheet name
                        original_category_name = clear_df[category_mask]['category'].iloc[0]
                        sheet_name = str(original_category_name)[:31]  # Limit to 31 chars
                        category_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        category_count += 1
                        
                        # Show category statistics
                        category_freq = category_df['frequency'].sum() if 'frequency' in category_df.columns else 0
                        print(f"  - {sheet_name}: {len(category_df)} entries, total frequency: {category_freq}")
                
                # Handle any remaining categories not in the fixed order
                processed_categories = [cat.lower() for cat in FIXED_CATEGORY_ORDER]
                remaining_categories = [
                    cat for cat in clear_df['category'].unique() 
                    if not pd.isna(cat) and str(cat).lower() not in processed_categories
                ]
                
                # Sort remaining categories alphabetically
                for category in sorted(remaining_categories):
                    if pd.isna(category) or category == '':
                        continue
                    
                    # Filter by category
                    category_df = clear_df[clear_df['category'] == category].copy()
                    
                    # Only create worksheet if there are entries
                    if not category_df.empty:
                        # Select columns using MAIN_COLUMNS
                        category_df = category_df[[c for c in MAIN_COLUMNS if c in category_df.columns]]
                        
                        # Limit sheet name to 31 characters (Excel limit)
                        sheet_name = str(category)[:31]
                        category_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        category_count += 1
                        
                        # Show category statistics
                        category_freq = category_df['frequency'].sum() if 'frequency' in category_df.columns else 0
                        print(f"  - {sheet_name}: {len(category_df)} entries, total frequency: {category_freq}")
                        
            elif not clear_df.empty:
                # If there's no category column or no categories, just write all clear entries
                clear_main_df = clear_df[[c for c in MAIN_COLUMNS if c in clear_df.columns]]
                clear_main_df.to_excel(writer, sheet_name='full glossary', index=False)
                category_count = 1
                print(f"  - Full Glossary: {len(clear_main_df)} entries")
            
            # For ambiguous entries: create a worksheet at the very end
            if not ambiguous_df.empty:
                ambiguous_main_df = ambiguous_df[[c for c in MAIN_COLUMNS if c in ambiguous_df.columns]]
                ambiguous_main_df.to_excel(writer, sheet_name='ambiguous terms', index=False)
                print(f"  - Ambiguous terms: {len(ambiguous_main_df)} entries")
            
            # Auto-adjust column widths for all sheets
            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✓ Main file created successfully with {category_count} category worksheet(s)")
        
        # 3. Create summary statistics (using fixed order for display)
        print(f"\n📊 Final Statistics:")
        print(f"   Total nouns: {len(df)}")
        
        if not clear_df.empty and 'category' in clear_df.columns:
            # Display statistics in fixed order first
            for category in FIXED_CATEGORY_ORDER:
                category_mask = clear_df['category'].str.lower() == category.lower()
                count = len(clear_df[category_mask])
                if count > 0:
                    freq_sum = clear_df[category_mask]['frequency'].sum()
                    original_name = clear_df[category_mask]['category'].iloc[0]
                    print(f"   - {original_name}: {count} entries, total frequency: {freq_sum}")
            
            # Then display remaining categories
            processed_categories = [cat.lower() for cat in FIXED_CATEGORY_ORDER]
            remaining_categories = [
                cat for cat in clear_df['category'].unique() 
                if not pd.isna(cat) and str(cat).lower() not in processed_categories
            ]
            
            for category in sorted(remaining_categories):
                count = len(clear_df[clear_df['category'] == category])
                if count > 0:
                    freq_sum = clear_df[clear_df['category'] == category]['frequency'].sum()
                    print(f"   - {category}: {count} entries, total frequency: {freq_sum}")
        
        if not ambiguous_df.empty:
            print(f"   - Ambiguous terms: {len(ambiguous_df)} entries")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False